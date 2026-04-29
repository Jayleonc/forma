"""Processor for Excel spreadsheet files (.xlsx / .xls)."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl.utils import get_column_letter

from .base import ExtractedVisualAsset, ProcessingResult, Processor

logger = logging.getLogger(__name__)


class XlsxProcessor(Processor):
    """Convert Excel spreadsheets to Markdown tables and extract embedded images."""

    _MAX_CONTEXT_FIELDS = 4
    _MAX_CONTEXT_VALUE_CHARS = 48

    def process(self, input_path: Path) -> ProcessingResult:
        import openpyxl

        logger.info("Processing xlsx file: %s", input_path)
        wb = openpyxl.load_workbook(input_path, read_only=False, data_only=True)

        parts: list[str] = []
        visual_assets: list[ExtractedVisualAsset] = []

        for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            sheet_parts = self._build_sheet_markdown(wb, sheet_name, rows)
            if sheet_parts:
                parts.extend(sheet_parts)
                parts.append("")

            visual_assets.extend(
                self._extract_sheet_images(
                    sheet_index=sheet_index,
                    sheet_name=sheet_name,
                    worksheet=ws,
                    rows=rows,
                )
            )

        wb.close()

        markdown = "\n".join(parts).strip()
        total_chars = len(markdown)
        logger.info(
            "Xlsx conversion complete: %d chars, %d sheets, %d images",
            total_chars,
            len(wb.sheetnames),
            len(visual_assets),
        )

        return ProcessingResult(
            markdown_content=markdown,
            text_char_count=total_chars,
            image_count=len(visual_assets),
            low_confidence=False,
            visual_assets=visual_assets,
        )

    def _build_sheet_markdown(self, workbook, sheet_name: str, rows: list[tuple[object, ...]]) -> list[str]:
        """Output each data row as an independent ### section.

        Each section contains the full header + that single row as a mini-table,
        so the downstream chunker treats every row as its own chunk.  This avoids
        many unrelated Q&A entries being merged into one embedding vector.
        """
        if not rows:
            return []

        parts: list[str] = []
        if len(workbook.sheetnames) > 1:
            parts.append(f"## {sheet_name}")
            parts.append("")

        max_col = self._max_populated_col(rows)
        if max_col == 0:
            return parts

        header = rows[0][:max_col]
        header_cells = [self._cell_str(c) for c in header]
        sep_cells = ["---"] * max_col

        for row_idx, row in enumerate(rows[1:], start=2):
            cells = [self._cell_str(row[i] if i < len(row) else None) for i in range(max_col)]

            # Skip completely empty rows
            if not any(c for c in cells):
                continue

            # Use first "meaningful" cell (len > 4) as the ### heading.
            # Short metadata values like "是"/"能"/"客服系统" are skipped so the
            # heading usually lands on the question/description column.
            title = (
                next((c for c in cells if len(c) > 4), None)
                or next((c for c in cells if c), f"行{row_idx}")
            )
            if len(title) > 60:
                title = title[:60]

            parts.append(f"### {title}")
            parts.append("")
            parts.append("| " + " | ".join(header_cells) + " |")
            parts.append("| " + " | ".join(sep_cells) + " |")
            parts.append("| " + " | ".join(cells) + " |")
            parts.append("")

        return parts

    def _extract_sheet_images(
        self,
        *,
        sheet_index: int,
        sheet_name: str,
        worksheet,
        rows: list[tuple[object, ...]],
    ) -> list[ExtractedVisualAsset]:
        images = list(getattr(worksheet, "_images", []))
        if not images:
            return []

        headers = self._row_headers(rows)
        assets: list[ExtractedVisualAsset] = []
        for idx, image in enumerate(images, start=1):
            anchor = self._extract_anchor(image.anchor)
            row_values = self._row_values(rows, anchor["from_row"])
            alt_text = self._build_alt_text(sheet_name, anchor, headers, row_values)
            extension = self._image_extension(image)
            mime_type = self._image_mime_type(extension)
            filename = f"{sheet_name}-r{anchor['from_row'] or 0}-c{anchor['from_col'] or 0}-{idx}.{extension}"
            assets.append(
                ExtractedVisualAsset(
                    filename=filename,
                    content=image._data(),
                    mime_type=mime_type,
                    alt_text=alt_text,
                    position_type="tabular_anchor",
                    position_meta={
                        "sheet": sheet_name,
                        "sheet_index": sheet_index,
                        "from_row": anchor["from_row"],
                        "to_row": anchor["to_row"],
                        "from_col": anchor["from_col"],
                        "to_col": anchor["to_col"],
                        "from_col_label": self._column_label(anchor["from_col"]),
                        "to_col_label": self._column_label(anchor["to_col"]),
                        "anchor_type": anchor["anchor_type"],
                        "context_text": self._build_row_context(headers, row_values),
                    },
                )
            )

        return assets

    def _build_alt_text(
        self,
        sheet_name: str,
        anchor: dict[str, int | str],
        headers: list[str],
        row_values: list[str],
    ) -> str:
        row_no = anchor["from_row"]
        col_no = anchor["from_col"]
        location = f'sheet "{sheet_name}"'
        if row_no:
            location += f", row {row_no}"
        if col_no:
            location += f", col {col_no}"

        context_parts: list[str] = []
        if row_no and row_no > 1 and headers and row_values:
            for header, value in zip(headers, row_values):
                clean_header = header.strip()
                clean_value = value.strip()
                if not clean_header or not clean_value:
                    continue
                context_parts.append(f"{clean_header}={clean_value}")
                if len(context_parts) >= self._MAX_CONTEXT_FIELDS:
                    break

        if context_parts:
            return f"Excel image at {location}; row context: {'; '.join(context_parts)}"
        return f"Excel image at {location}"

    def _build_row_context(self, headers: list[str], row_values: list[str]) -> str:
        context_parts: list[str] = []
        for header, value in zip(headers, row_values):
            clean_header = header.strip()
            clean_value = value.strip()
            if not clean_header or not clean_value:
                continue
            context_parts.append(f"{clean_header}={clean_value}")
            if len(context_parts) >= self._MAX_CONTEXT_FIELDS:
                break
        return "; ".join(context_parts)

    @staticmethod
    def _row_headers(rows: list[tuple[object, ...]]) -> list[str]:
        if not rows:
            return []
        max_col = XlsxProcessor._max_populated_col(rows)
        if max_col == 0:
            return []
        return [XlsxProcessor._cell_str(cell) for cell in rows[0][:max_col]]

    @staticmethod
    def _row_values(rows: list[tuple[object, ...]], row_number: int) -> list[str]:
        if row_number <= 0 or row_number > len(rows):
            return []
        row = rows[row_number - 1]
        return [XlsxProcessor._cell_str(cell)[: XlsxProcessor._MAX_CONTEXT_VALUE_CHARS] for cell in row]

    @staticmethod
    def _extract_anchor(anchor) -> dict[str, int | str]:
        from_marker = getattr(anchor, "_from", None)
        to_marker = getattr(anchor, "to", None)
        anchor_type = anchor.__class__.__name__

        from_row = XlsxProcessor._marker_value(from_marker, "row")
        from_col = XlsxProcessor._marker_value(from_marker, "col")
        to_row = XlsxProcessor._marker_value(to_marker, "row")
        to_col = XlsxProcessor._marker_value(to_marker, "col")

        if from_row == 0 and hasattr(anchor, "row"):
            from_row = int(getattr(anchor, "row", 0))
        if from_col == 0 and hasattr(anchor, "col"):
            from_col = int(getattr(anchor, "col", 0))

        if to_row == 0:
            to_row = from_row
        if to_col == 0:
            to_col = from_col

        return {
            "from_row": from_row,
            "to_row": to_row,
            "from_col": from_col,
            "to_col": to_col,
            "anchor_type": anchor_type,
        }

    @staticmethod
    def _marker_value(marker, field_name: str) -> int:
        if marker is None:
            return 0
        raw = int(getattr(marker, field_name, 0) or 0)
        return raw + 1 if raw >= 0 else 0

    @staticmethod
    def _max_populated_col(rows: list[tuple[object, ...]]) -> int:
        max_col = 0
        for row in rows:
            for i in range(len(row) - 1, -1, -1):
                if row[i] is not None:
                    max_col = max(max_col, i + 1)
                    break
        return max_col

    @staticmethod
    def _image_extension(image) -> str:
        image_format = (getattr(image, "format", "") or "").lower()
        if image_format in {"png", "jpeg", "jpg", "gif"}:
            return "jpg" if image_format == "jpeg" else image_format
        return "png"

    @staticmethod
    def _image_mime_type(extension: str) -> str:
        if extension == "jpg":
            return "image/jpeg"
        if extension == "gif":
            return "image/gif"
        return "image/png"

    @staticmethod
    def _column_label(column_number: int) -> str:
        if column_number <= 0:
            return ""
        return get_column_letter(column_number)

    @staticmethod
    def _cell_str(value: object) -> str:
        if value is None:
            return ""
        s = str(value).replace("|", "\\|").replace("\n", " ")
        return s.strip()
